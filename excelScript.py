import openpyxl as py

wb = py.load_workbook(r'C:\\\Users\Laviz Pandey\Documents\Form.xlsx')
sh = wb['Sheet1']
row = sh.max_row
column = sh.max_column

maillist = []
citylist = []


def excelClass(cntrLocation):
    for i in range(1, row + 1):
        countryfromList = sh.cell(i, 3).value
        if countryfromList == cntrLocation:
            mailValue = sh.cell(i, 4).value
            maillist.append(mailValue)


for i in range(1, row + 1):
    data = sh.cell(i, 3).value
    cityValue = sh.cell(i, 3).value
    citylist.append(cityValue)
